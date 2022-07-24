import openpyxl

def getTestData():
    list = []
    workBook = openpyxl.load_workbook("C:/Users/vinay/PycharmProjects/Assignment001/TestData/TestDataSheet.xlsx")
    sheet = workBook.active
    rows = sheet.max_row

    for r in range(2, rows+1):

        UserId = sheet.cell(r,2).value
        Password = sheet.cell(r,3).value
        E_Condition = sheet.cell(r,4).value

        tuple = (UserId, Password, E_Condition)
        list.append(tuple)

    return list
